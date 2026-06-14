from app.models.user import UserCreate, RegisterRequest, UserBindInfoRequest
from app.models.user import ImportTemplateRequest, ImportTemplateResponse, ImportResult
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from app.schemas import User, ImportTemplate, UserCredential, UserTenant
from fastapi import HTTPException
import json
import base64
import io
from openpyxl import load_workbook

from app.crud import crud_credential
from app.core.pii import blind_index, normalize_optional_text


def _is_valid_avatar_url(value: str) -> bool:
    text = (value or "").strip()
    if not text:
        return False
    if text in {
        "builtin:avatar-1",
        "builtin:avatar-2",
        "builtin:avatar-3",
        "builtin:avatar-4",
    }:
        return True
    if text.startswith("/uploads/avatars/"):
        return True
    if text.startswith("http://") or text.startswith("https://"):
        return "/uploads/avatars/" in text
    return False


def get_users(db: Session, tenant_id: int) -> list[User]:
    """获取用户列表（租户隔离）"""
    return db.query(User).filter(User.tenant_id == tenant_id).all()


def get_user(db: Session, user_id: int, tenant_id: int) -> User | None:
    """获取单个用户（租户隔离）"""
    return db.query(User).filter(
        User.id == user_id,
        User.tenant_id == tenant_id
    ).first()


def get_user_by_phone(db: Session, phone: str, tenant_id: int) -> User | None:
    """根据手机号获取用户（租户隔离）"""
    phone_hash = blind_index(phone, purpose="phone")
    if not phone_hash:
        return None
    return db.query(User).filter(
        User.phone_hash == phone_hash,
        User.tenant_id == tenant_id
    ).first()


def get_user_by_wx_openid(db: Session, openid: str, tenant_id: int) -> User | None:
    """根据微信 openid 获取用户（通过 user_credential）。"""
    return db.query(User).join(
        UserCredential,
        UserCredential.user_id == User.id,
    ).filter(
        UserCredential.tenant_id == tenant_id,
        UserCredential.credential_type == "wechat",
        UserCredential.identifier == openid,
        UserCredential.status == 1,
        User.tenant_id == tenant_id,
    ).first()


def _ensure_user_tenant_membership(db: Session, user_id: int, tenant_id: int) -> None:
    membership = db.query(UserTenant).filter(
        UserTenant.user_id == user_id,
        UserTenant.tenant_id == tenant_id,
    ).first()
    if membership:
        if membership.status != 1:
            membership.status = 1
        return
    db.add(UserTenant(user_id=user_id, tenant_id=tenant_id, status=1))
    db.flush()


def get_or_create_user_wechat(db: Session, openid: str, tenant_id: int, nickname: str | None = None) -> User:
    """微信授权登录：存在则返回，不存在则创建"""
    user = get_user_by_wx_openid(db, openid, tenant_id)
    if user:
        _ensure_user_tenant_membership(db, user.id, tenant_id)
        return user
    try:
        db_user = User(
            tenant_id=tenant_id,
            name=nickname or "微信用户",
            phone=None,
            email=None,
            identity_number=None,
            sex=None,
            isblock=0,
            block_reason=None,
        )
        db.add(db_user)
        db.flush()
        _ensure_user_tenant_membership(db, db_user.id, tenant_id)
        crud_credential.bind_wechat_credential(
            db,
            user_id=db_user.id,
            tenant_id=tenant_id,
            openid=openid,
        )
        db.commit()
        db.refresh(db_user)
        return db_user
    except IntegrityError:
        db.rollback()
        # 竞态条件：可能其他请求已创建，重新查询
        user = get_user_by_wx_openid(db, openid, tenant_id)
        if user:
            return user
        raise HTTPException(status_code=400, detail=f"创建微信用户失败")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"创建微信用户失败: {str(e)}")


def create_user(db: Session, user: UserCreate, tenant_id: int) -> User:
    """创建用户（租户隔离）"""
    try:
        if user.identity_number:
            existing = db.query(User).filter(
                User.identity_number_hash == blind_index(user.identity_number, purpose="identity_number"),
                User.tenant_id == tenant_id
            ).first()
            if existing:
                raise HTTPException(status_code=400, detail="该证件号已存在")

        db_user = User(
            tenant_id=tenant_id,
            name=user.name,
            identity_number=user.identity_number,
            phone=user.phone,
            email=user.email,
            sex=user.sex,
            isblock=user.isblock,
            block_reason=user.block_reason,
        )
        db.add(db_user)
        db.flush()
        _ensure_user_tenant_membership(db, db_user.id, tenant_id)
        db.commit()
        db.refresh(db_user)
        return db_user
    except HTTPException:
        raise
    except IntegrityError:
        db.rollback()
        # 竞态条件：可能是手机号或邮箱重复
        if user.phone:
            existing = get_user_by_phone(db, user.phone, tenant_id)
            if existing:
                raise HTTPException(status_code=400, detail="该手机号已存在")
        if user.email:
            existing = db.query(User).filter(
                User.email_hash == blind_index(user.email, purpose="email"),
                User.tenant_id == tenant_id
            ).first()
            if existing:
                raise HTTPException(status_code=400, detail="该邮箱已存在")
        raise HTTPException(status_code=400, detail="创建用户失败")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


def register_user(db: Session, body: RegisterRequest, tenant_id: int) -> User:
    """用户注册（租户隔离）"""
    try:
        existing_phone = get_user_by_phone(db, body.phone.strip(), tenant_id)
        if existing_phone:
            raise HTTPException(status_code=400, detail="该手机号已注册")

        if body.email:
            existing_email = db.query(User).filter(
                User.email_hash == blind_index(body.email, purpose="email"),
                User.tenant_id == tenant_id
            ).first()
            if existing_email:
                raise HTTPException(status_code=400, detail="该邮箱已注册")

        db_user = User(
            tenant_id=tenant_id,
            name=body.name.strip(),
            phone=body.phone.strip(),
            email=body.email,
            identity_number=None,
            sex=None,
            isblock=0,
            block_reason=None,
        )
        db.add(db_user)
        db.flush()
        _ensure_user_tenant_membership(db, db_user.id, tenant_id)
        crud_credential.create_password_credential(
            db=db,
            user_id=db_user.id,
            tenant_id=tenant_id,
            identifier=body.phone.strip(),
            password=body.password,
            must_reset=False,
        )
        db.commit()
        db.refresh(db_user)
        return db_user
    except HTTPException:
        raise
    except IntegrityError:
        db.rollback()
        # 竞态条件：可能是手机号或邮箱重复
        existing_phone = get_user_by_phone(db, body.phone.strip(), tenant_id)
        if existing_phone:
            raise HTTPException(status_code=400, detail="该手机号已注册")
        if body.email:
            existing_email = db.query(User).filter(
                User.email_hash == blind_index(body.email, purpose="email"),
                User.tenant_id == tenant_id
            ).first()
            if existing_email:
                raise HTTPException(status_code=400, detail="该邮箱已注册")
        raise HTTPException(status_code=400, detail="注册失败")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


def authenticate_user(db: Session, phone: str, password: str, tenant_id: int) -> User | None:
    """普通用户认证（通过 user_credential 进行密码校验）。"""
    cred = crud_credential.authenticate_by_password(db, tenant_id, phone, password)
    if not cred:
        return None
    return get_user(db, cred.user_id, tenant_id)


def is_user_profile_incomplete(db: Session, user_id: int, tenant_id: int) -> bool:
    """
    检查用户信息是否不完整。
    判断依据是实际字段的有无，不依赖默认昵称的字符串值，
    避免因默认昵称变更导致的静默失效。
    """
    user = get_user(db, user_id, tenant_id)
    if not user:
        return True
    # 姓名必须存在且不为空
    if not user.name or not user.name.strip():
        return True
    if not user.sex:
        return True
    # 手机号必须存在且不是系统生成的占位符
    if not user.phone or user.phone.startswith("wx_"):
        return True
    if user.age is None:
        return True
    if not user.occupation or not str(user.occupation).strip():
        return True
    if not user.industry or not str(user.industry).strip():
        return True
    return False


def update_user_bind_info(db: Session, user_id: int, tenant_id: int, bind_info: UserBindInfoRequest) -> User:
    """更新用户绑定信息（不再采集证件信息）。"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 性别格式转换：将 male/female 统一转为数据库存储格式 M/F
    sex_map = {"male": "M", "female": "F"}
    if bind_info.sex not in sex_map:
        raise HTTPException(status_code=400, detail="请选择男或女")
    sex_value = sex_map[bind_info.sex]

    # 验证手机号一致性：用户填写的手机号必须与微信获取的手机号一致
    bind_phone = bind_info.phone
    current_phone = user.phone
    # 如果当前手机号是有效的（不是 wx_ 开头的占位符）
    if current_phone and not current_phone.startswith("wx_"):
        if bind_phone != current_phone:
            raise HTTPException(
                status_code=400,
                detail=f"手机号不一致，请使用微信授权的手机号 {current_phone[:3]}****{current_phone[-4:]}"
            )
    # 检查手机号是否被其他用户占用
    existing = db.query(User).filter(
        User.phone_hash == blind_index(bind_phone, purpose="phone"),
        User.id != user_id,
        User.tenant_id == tenant_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该手机号已被使用")

    old_phone = user.phone
    avatar_url = (bind_info.avatar_url or "").strip()
    if avatar_url and not _is_valid_avatar_url(avatar_url):
        raise HTTPException(status_code=400, detail="头像地址不合法")

    # 更新字段
    user.name = bind_info.name
    user.sex = sex_value
    user.age = bind_info.age
    user.occupation = bind_info.occupation
    user.phone = bind_phone
    user.email = bind_info.email
    user.industry = bind_info.industry
    if avatar_url:
        user.avatar_url = avatar_url

    crud_credential.sync_phone_identifiers(db, user_id, tenant_id, old_phone, bind_phone)
    db.commit()
    db.refresh(user)
    return user


def update_user_avatar(db: Session, user_id: int, tenant_id: int, avatar_url: str) -> User:
    """更新用户头像。"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    normalized_avatar_url = (avatar_url or "").strip()
    if not _is_valid_avatar_url(normalized_avatar_url):
        raise HTTPException(status_code=400, detail="头像地址不合法")

    user.avatar_url = normalized_avatar_url
    db.commit()
    db.refresh(user)
    return user


def update_user_profile_self(
    db: Session,
    user_id: int,
    tenant_id: int,
    *,
    name: str,
    sex: str,
    age: int,
    occupation: str,
    industry: str,
    email: str | None,
) -> User:
    """用户自主修改资料（不含手机号/证件号）。"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    sex_map = {"male": "M", "female": "F"}
    if sex not in sex_map:
        raise HTTPException(status_code=400, detail="请选择男或女")

    if email:
        existing_email = db.query(User).filter(
            User.email_hash == blind_index(email, purpose="email"),
            User.tenant_id == tenant_id,
            User.id != user_id,
        ).first()
        if existing_email:
            raise HTTPException(status_code=400, detail="该邮箱已被使用")

    user.name = name.strip()
    user.sex = sex_map[sex]
    user.age = age
    user.occupation = occupation.strip()
    user.industry = industry.strip()
    user.email = email
    db.commit()
    db.refresh(user)
    return user


def clear_user_profile_fields_self(
    db: Session,
    user_id: int,
    tenant_id: int,
    fields: list[str],
) -> User:
    """用户自主删除部分个人信息字段（不含身份证相关字段）。"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    for field_name in fields:
        setattr(user, field_name, None)
    db.commit()
    db.refresh(user)
    return user


def deactivate_user_account(
    db: Session,
    user_id: int,
    tenant_id: int,
) -> User:
    """注销账号：清空可删除资料并拉黑账号，防止继续登录。"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    user.name = "已注销用户"
    user.sex = None
    user.age = None
    user.occupation = None
    user.industry = None
    user.email = None
    user.avatar_url = None
    user.isblock = 1
    user.block_reason = "用户自主注销"
    db.flush()
    return user


def get_or_create_user_by_phone(db: Session, phone: str, tenant_id: int, name: str | None = None) -> User:
    """根据手机号查找或创建用户（用于手机号授权登录）"""
    user = get_user_by_phone(db, phone, tenant_id)
    if user:
        _ensure_user_tenant_membership(db, user.id, tenant_id)
        return user
    try:
        db_user = User(
            tenant_id=tenant_id,
            name=name or f"用户{phone[-4:]}",
            phone=phone,
            email=None,
            identity_number=None,
            sex=None,
            isblock=0,
            block_reason=None,
        )
        db.add(db_user)
        db.flush()
        _ensure_user_tenant_membership(db, db_user.id, tenant_id)
        db.commit()
        db.refresh(db_user)
        return db_user
    except IntegrityError:
        db.rollback()
        # 竞态条件：可能其他请求已创建，重新查询
        user = get_user_by_phone(db, phone, tenant_id)
        if user:
            return user
        raise HTTPException(status_code=400, detail="创建用户失败")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"创建用户失败: {str(e)}")


def get_all_users_for_super_admin(
    db: Session,
    tenant_id: int,
    skip: int = 0,
    limit: int = 20,
    keyword: str | None = None,
) -> tuple[list[User], int]:
    """
    超级管理员查看指定租户的用户（支持分页和搜索）
    返回 (用户列表, 总数)
    """
    users = db.query(User).filter(User.tenant_id == tenant_id).order_by(User.id.desc()).all()

    if keyword:
        normalized_keyword = normalize_optional_text(keyword)
        phone_hash = blind_index(normalized_keyword, purpose="phone")
        keyword_lower = (normalized_keyword or "").lower()

        def _matches(item: User) -> bool:
            name = (item.name or "").lower()
            phone = item.phone or ""
            return (
                (keyword_lower and keyword_lower in name)
                or (phone_hash is not None and item.phone_hash == phone_hash)
                or (keyword_lower and keyword_lower in phone)
            )

        users = [item for item in users if _matches(item)]

    total = len(users)
    paged_users = users[skip: skip + limit]
    return paged_users, total


def block_user(db: Session, user_id: int, tenant_id: int, reason: str | None = None) -> User:
    """拉黑用户"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if user.isblock == 1:
        raise HTTPException(status_code=400, detail="用户已被拉黑")
    user.isblock = 1
    user.block_reason = reason
    db.commit()
    db.refresh(user)
    return user


def unblock_user(db: Session, user_id: int, tenant_id: int) -> User:
    """解除拉黑用户"""
    user = get_user(db, user_id, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if user.isblock == 0:
        raise HTTPException(status_code=400, detail="用户未被拉黑")
    user.isblock = 0
    user.block_reason = None
    db.commit()
    db.refresh(user)
    return user


# ============================================================
# 导入模板配置 CRUD
# ============================================================

def get_import_template(db: Session, tenant_id: int) -> ImportTemplate | None:
    """获取导入模板配置"""
    return db.query(ImportTemplate).filter(ImportTemplate.tenant_id == tenant_id).first()


def save_import_template(db: Session, tenant_id: int, config: ImportTemplateRequest) -> ImportTemplate:
    """保存导入模板配置"""
    template = get_import_template(db, tenant_id)
    if template:
        template.column_mapping = json.dumps(config.column_mapping)
        template.is_active = 1
    else:
        template = ImportTemplate(
            tenant_id=tenant_id,
            column_mapping=json.dumps(config.column_mapping),
            is_active=1,
        )
        db.add(template)
    db.commit()
    db.refresh(template)
    return template


def delete_import_template(db: Session, tenant_id: int) -> bool:
    """删除导入模板配置"""
    template = get_import_template(db, tenant_id)
    if not template:
        return False
    template.is_active = 0
    db.commit()
    return True


# 用户字段映射
USER_FIELD_LABELS = {
    "name": "姓名",
    "sex": "性别",
    "phone": "手机号",
    "email": "邮箱",
    "identity_type": "证件类型",
    "identity_number": "证件号码",
    "occupation": "职业",
    "industry": "行业",
    "age": "年龄",
    "isblock": "是否拉黑",
    "block_reason": "拉黑原因",
}


def _normalize_import_identity_type(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    mapping = {
        "mainland": "mainland",
        "大陆身份证": "mainland",
        "大陆": "mainland",
        "身份证": "mainland",
        "中国大陆身份证": "mainland",
        "hongkong": "hongkong",
        "港澳台通行证": "hongkong",
        "港澳台证件": "hongkong",
        "香港证件": "hongkong",
        "香港": "hongkong",
        "foreign": "foreign",
        "护照": "foreign",
    }
    return mapping.get(text, text)


def _normalize_import_block_value(value: str | None) -> int:
    if not value:
        return 0
    text = value.strip().lower()
    if text in {"1", "true", "yes", "y", "是", "拉黑", "黑名单", "已拉黑"}:
        return 1
    return 0


def import_users_from_excel(db: Session, tenant_id: int, file_content_b64: str) -> ImportResult:
    """
    从Excel文件导入用户
    第一行是表头，不导入数据
    """
    template = get_import_template(db, tenant_id)
    if not template or not template.column_mapping:
        raise HTTPException(status_code=400, detail="请先配置导入模板")

    column_mapping = json.loads(template.column_mapping)

    # 解码Base64文件内容
    file_content = base64.b64decode(file_content_b64)
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    success_count = 0
    failed_count = 0
    errors = []

    # 性别映射
    sex_map = {"男": "M", "女": "F", "M": "M", "F": "F"}

    # 从第二行开始读取（跳过表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            user_data = {}
            for col_idx, field_name in column_mapping.items():
                cell_value = row[int(col_idx)] if int(col_idx) < len(row) else None
                if cell_value is not None:
                    cell_value = str(cell_value).strip()

                # 性别转换
                if field_name == "sex" and cell_value:
                    cell_value = sex_map.get(cell_value, cell_value)
                if field_name == "identity_type":
                    cell_value = _normalize_import_identity_type(cell_value)

                if cell_value:
                    user_data[field_name] = cell_value

            # 至少需要有姓名
            if not user_data.get("name"):
                failed_count += 1
                errors.append(f"第{row_idx}行：姓名为空")
                continue

            # 创建用户
            db_user = User(
                tenant_id=tenant_id,
                name=user_data.get("name"),
                phone=user_data.get("phone"),
                email=user_data.get("email"),
                sex=user_data.get("sex"),
                age=int(user_data["age"]) if user_data.get("age") else None,
                occupation=user_data.get("occupation"),
                industry=user_data.get("industry"),
                identity_type=user_data.get("identity_type"),
                identity_number=user_data.get("identity_number"),
                isblock=_normalize_import_block_value(user_data.get("isblock")),
                block_reason=user_data.get("block_reason"),
            )
            db.add(db_user)
            db.commit()
            success_count += 1

        except IntegrityError as e:
            db.rollback()
            failed_count += 1
            error_msg = str(e.orig) if hasattr(e, 'orig') else str(e)
            if "phone" in error_msg.lower() or "duplicate" in error_msg.lower():
                errors.append(f"第{row_idx}行：手机号已存在")
            elif "identity" in error_msg.lower():
                errors.append(f"第{row_idx}行：证件号已存在")
            else:
                errors.append(f"第{row_idx}行：数据冲突")
        except Exception as e:
            db.rollback()
            failed_count += 1
            errors.append(f"第{row_idx}行：{str(e)[:50]}")

    return ImportResult(success=success_count, failed=failed_count, errors=errors[:20])
